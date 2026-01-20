import streamlit as st
from supabase import create_client
import json
import pandas as pd
from datetime import datetime


st.set_page_config(page_title="Prompt Manager", page_icon="📋", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebarNav"] { display: none; }
button { height: 42px; }
</style>
""", unsafe_allow_html=True)


if "user_email" not in st.session_state:
    st.switch_page("login.py")
    st.stop()

ADMIN = st.session_state.get("user_role") == "admin"


@st.cache_resource
def get_supabase():
    return create_client(
        st.secrets["SUPABASE_URL"],
        st.secrets["SUPABASE_ANON_KEY"]
    )

supabase = get_supabase()


st.title("📋 Prompt Manager")

col1, col2 = st.columns([4, 1])
with col1:
    st.caption(f"Logged in as **{st.session_state.user_email}**")
with col2:
    if st.button("🔄 Refresh", use_container_width=True):
        st.cache_data.clear()
        st.session_state["_force_refresh"] = True
        st.rerun()


@st.cache_data(ttl=60)
def fetch_prompts(_force=None):
    return supabase.table("prompts").select("*").order(
        "created_at", desc=True
    ).execute().data

@st.cache_data(ttl=60)
def fetch_ratings(_force=None):
    return supabase.table("ratings").select("*").execute().data

@st.cache_data(ttl=60)
def fetch_notes(_force=None):
    return supabase.table("notes").select("*").order(
        "created_at", desc=False
    ).execute().data

prompts = fetch_prompts(st.session_state.get("_force_refresh"))
ratings = fetch_ratings(st.session_state.get("_force_refresh"))
notes = fetch_notes(st.session_state.get("_force_refresh"))


if "_force_refresh" in st.session_state:
    del st.session_state["_force_refresh"]


rating_map = {
    (r["prompt_id"], r["user_email"]): r["rating"]
    for r in ratings
}


notes_map = {}
for note in notes:
    pid = note["prompt_id"]
    if pid not in notes_map:
        notes_map[pid] = []
    notes_map[pid].append(note)


with st.expander("📁 Bulk Import from JSON"):
    file = st.file_uploader("Upload JSON", type=["json"])
    if file:
        try:
            data = json.load(file)
            st.write(f"Found **{len(data)}** entries")

            if st.button("Import All"):
                for d in data:
                    supabase.table("prompts").insert({
                        "prompt": d.get("prompt", ""),
                        "query": d.get("query", ""),
                        "response": d.get("response", ""),
                        "created_by": st.session_state.user_email
                    }).execute()
                st.success("Import completed")
                st.cache_data.clear()
                st.session_state["_force_refresh"] = True
                st.rerun()
        except Exception as e:
            st.error(str(e))

if prompts:
    with st.expander("📤 Export Data"):

        
        export_rows = []
        for p in prompts:
            export_rows.append({
                "prompt": p["prompt"],
                "query": p["query"],
                "response": p["response"],
                "rating": rating_map.get((p["id"], st.session_state.user_email))  
            })

        df = pd.DataFrame(export_rows)

        col1, col2, col3 = st.columns(3)

        
        with col1:
            st.download_button(
                "⬇️ Download JSON",
                json.dumps(
                    [{k: v for k, v in row.items() if k != "rating"} for row in export_rows],
                    indent=2
                ),
                f"prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
                "application/json",
                use_container_width=True
            )

        
        with col2:
            st.download_button(
                "⬇️ Download CSV",
                df.drop(columns=["rating"]).to_csv(index=False),
                f"prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                "text/csv",
                use_container_width=True
            )

       
        with col3:
            import openpyxl
            from openpyxl.styles import PatternFill

            xlsx_path = f"/tmp/prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                sheet_df = df.drop(columns=["rating"])
                sheet_df.to_excel(writer, index=False, sheet_name="prompts")

                worksheet = writer.sheets["prompts"]

                GREEN = "C6EFCE"
                RED = "FFC7CE"

                for i, row in enumerate(export_rows, start=2): 
                    rating = row["rating"]

                    if rating == "up":
                        fill_color = GREEN
                    elif rating == "down":
                        fill_color = RED
                    else:
                        continue

                    for col in range(1, 4):  
                        cell = worksheet.cell(row=i, column=col)
                        cell.fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type="solid"
                        )

            with open(xlsx_path, "rb") as f:
                st.download_button(
                    "⬇️ Download XLSX",
                    f,
                    file_name=f"prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with open(xlsx_path, "rb") as f:
                st.download_button(
                    "⬇️ Download XLSX",
                    f,
                    file_name=f"prompts_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )



with st.expander("➕ Add New Prompt"):
    with st.form("add_prompt", clear_on_submit=True):
        p = st.text_input("Prompt")
        q = st.text_area("Query", height=120)
        r = st.text_area("Response", height=120)

        if st.form_submit_button("💾 Save"):
            if p and q and r:
                supabase.table("prompts").insert({
                    "prompt": p,
                    "query": q,
                    "response": r,
                    "created_by": st.session_state.user_email
                }).execute()
                st.success("Prompt added")
                st.cache_data.clear()
                st.session_state["_force_refresh"] = True
                st.rerun()
            else:
                st.error("All fields required")


col1, col2, col3 = st.columns([3, 2, 2])

with col1:
    search = st.text_input("🔍 Search")

with col2:
    sort = st.selectbox("Sort", ["Newest First", "Oldest First"])

with col3:
    rating_filter = st.selectbox(
        "Filter by rating",
        ["All", "👍 Liked", "👎 Disliked", "Unrated"]
    )


def get_user_rating(prompt_id):
    return rating_map.get((prompt_id, st.session_state.user_email))

filtered = []

for p in prompts:
    rating = get_user_rating(p["id"])

    if rating_filter == "👍 Liked" and rating != "up":
        continue
    if rating_filter == "👎 Disliked" and rating != "down":
        continue
    if rating_filter == "Unrated" and rating is not None:
        continue

    if search:
        text = f"{p['prompt']} {p['query']} {p['response']}".lower()
        if search.lower() not in text:
            continue

    filtered.append(p)

if sort == "Oldest First":
    filtered.reverse()


st.divider()
st.markdown(f"### 📚 Prompts ({len(filtered)})")

if not filtered:
    st.info("No prompts found")
    st.stop()

for idx, item in enumerate(filtered):
    rating = get_user_rating(item["id"])

    h, up, down, delete = st.columns([6, 1, 1, 1])

    with h:
       
        rating_indicator = ""
        if rating == "up":
            rating_indicator = " 👍"
        elif rating == "down":
            rating_indicator = " 👎"
        st.markdown(f"**{idx + 1}. {item['prompt']}**{rating_indicator}")

    with up:
        if st.button("👍", key=f"up_{item['id']}"):
            supabase.table("ratings").upsert({
                "prompt_id": item["id"],
                "user_email": st.session_state.user_email,
                "rating": "up"
            }, on_conflict="prompt_id,user_email").execute()
            st.cache_data.clear()
            st.session_state["_force_refresh"] = True
            st.rerun()

    with down:
        if st.button("👎", key=f"down_{item['id']}"):
            supabase.table("ratings").upsert({
                "prompt_id": item["id"],
                "user_email": st.session_state.user_email,
                "rating": "down"
            }, on_conflict="prompt_id,user_email").execute()
            st.cache_data.clear()
            st.session_state["_force_refresh"] = True
            st.rerun()

    with delete:
        if ADMIN and st.button("🗑️", key=f"del_{item['id']}"):
           
            supabase.table("ratings").delete().eq("prompt_id", item["id"]).execute()
            supabase.table("notes").delete().eq("prompt_id", item["id"]).execute()
            supabase.table("prompts").delete().eq("id", item["id"]).execute()
            st.cache_data.clear()
            st.session_state["_force_refresh"] = True
            st.rerun()

    with st.expander("View details"):
     
        edit_key = f"edit_{item['id']}"
        if edit_key not in st.session_state:
            st.session_state[edit_key] = False

      
        if st.button(
            "✏️ Edit" if not st.session_state[edit_key] else "❌ Cancel Edit",
            key=f"btn_{item['id']}"
        ):
            st.session_state[edit_key] = not st.session_state[edit_key]
            st.rerun()

       
        if st.session_state[edit_key]:
            
            with st.form(f"edit_form_{item['id']}"):
                edited_prompt = st.text_input("Prompt", value=item["prompt"])
                edited_query = st.text_area("Query", value=item["query"], height=120)
                edited_response = st.text_area("Response", value=item["response"], height=120)

                if st.form_submit_button("💾 Update"):
                    if edited_prompt and edited_query and edited_response:
                        supabase.table("prompts").update({
                            "prompt": edited_prompt,
                            "query": edited_query,
                            "response": edited_response,
                            "updated_at": datetime.utcnow().isoformat(),
                            "last_modified_by": st.session_state.user_email
                        }).eq("id", item["id"]).execute()

                        st.session_state[edit_key] = False
                        st.cache_data.clear()
                        st.session_state["_force_refresh"] = True
                        st.success("Updated successfully")
                        st.rerun()
                    else:
                        st.error("All fields required")
        else:
          
            st.markdown("**Prompt:**")
            st.write(item["prompt"])

            st.markdown("**Query:**")
            st.code(item["query"], language="sql")

            st.markdown("**Response:**")
            st.write(item["response"])

       
        metadata_text = f"Created by {item.get('created_by', 'Unknown')} | {item.get('created_at', 'N/A')}"

        if item.get('last_modified_by'):
            metadata_text += f"\n\nLast modified by **{item['last_modified_by']}** | {item.get('updated_at', 'N/A')}"

        st.caption(metadata_text)

       
        item_notes = notes_map.get(item["id"], [])
        if item_notes:
            st.markdown("**Notes:**")
            for note in item_notes:
                st.info(
                    f"💬 {note['note']}\n\n"
                    f"*— {note.get('created_by', 'Unknown')} | "
                    f"{note.get('created_at', 'N/A')}*"
                )

      
        with st.form(f"note_{item['id']}"):
            note = st.text_input("Add note")
            if st.form_submit_button("💬 Save Note"):
                if note:
                    supabase.table("notes").insert({
                        "prompt_id": item["id"],
                        "note": note,
                        "created_by": st.session_state.user_email
                    }).execute()
                    st.success("Note added")
                    st.cache_data.clear()
                    st.session_state["_force_refresh"] = True
                    st.rerun()
